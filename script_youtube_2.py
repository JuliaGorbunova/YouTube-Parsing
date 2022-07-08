from selenium import webdriver
from bs4 import BeautifulSoup as BS
import time
import os.path
import xlwt
import openpyxl
from openpyxl import load_workbook
import sys

from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from data import channels
from auth import authorisation
from proxy import proxy_chrome
import csv
driver = webdriver.Chrome()
base_url = "https://www.youtube.com/"
link_of_video=''

#     инициализируем списки логинов, паролей (ютуб), прокси-логинов, прокси-паролей, хостов и портов
# youtube_logins=[]
# youtube_passw=[]
# proxy_logins=[]
# proxy_pass=[]
# proxy_hosts=[]
# proxy_ports=[]

# открываем файл с логинами,паролями и прокси-серверами, читаем оттуда данные и заполняем списки
# with open("auth.csv", encoding='utf-8') as r_file:
#     reader_object = csv.reader(r_file, delimiter = ";")
#     for row in reader_object:
#         youtube_logins.append(row[0])
#         youtube_passw.append(row[1])
#         proxy_logins.append(row[2][0:row[2].index(':')])
#         proxy_pass.append(row[2][row[2].index(':')+1:row[2].index('@')])
#         proxy_hosts.append(row[2][row[2].index('@')+1:row[2].index(':')])
#         proxy_ports.append(int((row[2][len(row[2])-4:])))
# print (log_pass)

# driver = proxy_chrome(PROXY_HOST,PROXY_PORT,PROXY_USER,PROXY_PASS)

driver.set_window_size(1920, 1080)


def get_info_in_main_page(quantity,driver):
    global link_of_video
    driver.get(base_url)
    time.sleep(10)
    # получаем текст исходного кода страницы
    html = driver.page_source
    # извлекаем данные о всех названиях видео из кода страницы
    soup = BS(html, "html.parser")
    list_of_videos=soup.find_all("a",{"class":"yt-simple-endpoint focus-on-expand style-scope ytd-rich-grid-media"})
    list_of_channels_all=soup.find_all("a",{"class":"yt-simple-endpoint style-scope yt-formatted-string"})
    # создаем пустые списки, куда в дальнейшем поместим имена видео, ссылки на них, а также названия каналов
    list_of_videos_names=[]
    list_of_links = []
    list_of_channels=[]
    # перебираем значения полученного массива видео, достаем оттуда имена и ссылки, помещаем их в новые списки
    for video in list_of_videos[0:20]:
        list_of_links.append(base_url+video.get('href'))
        list_of_videos_names.append(video.get('title'))
    #  перебираем значения полученного массива каналов, достаем оттуда ссылки на каналы
    for item in list_of_channels_all[0:20]:
        list_of_channels.append(item.get_text())
    # дальше работа с excel. Проверим, есть ли в директории файл. Если нет, создадим
    if not os.path.isfile('table_of_videos.xlsx'):
        # Создаем таблицу (рабочую книгу)
        table_of_videos=openpyxl.Workbook()
        sheet=table_of_videos.active
        # делаем заглавие для первого листа
        table_of_videos.create_sheet(title='Статистика',index=0)
        # далее работаем на первом листе
        table_of_videos.save('table_of_videos.xlsx')
        sheet=table_of_videos['Статистика']
        # забиваем заголовки столбцов
        sheet['A1']='Аккаунт (рандомный айди)'
        sheet['B1']='Сколько видео из списка А1 посмотрел'
        sheet['C1']='Рекламный блок'
        sheet['D1']='Позиция'
        sheet['E1']='Ссылка на последнее просмотренное видео'
        sheet['F1']='Название'
        sheet['G1']='URL видео'
        sheet['H1']='URL канала'
    #     если файл есть, откроем его
    else:
        table_of_videos = load_workbook('table_of_videos.xlsx')
    # ищем первую пустую строку,чтобы продолжить заполнять таблицу
    sheet = table_of_videos.active
    index=sheet.max_row+1
    #     заполняем 20 строк таблицы данными, которые получили в ходе парсинга
    # забиваем значения первых двадцати строк
    # заводим счетчик для того, чтобы брать элементы из списков имен, ссылок и каналов
    j=0
    # еще один счетчик для заполнения столбца "Позиция"
    position=1
    for i in range(index,index+20):
        sheet[f'A{i}'] = 'random'
        sheet[f'B{i}']=quantity
        sheet[f'C{i}']='главная страница'
        sheet[f'D{i}']=position
        # проверяем, есть ли какое-то значение в переменной link_of_video
        # если парсинг с главной выполняется впервые, значения не будет
        # если уже был заход в канал, значение будет, оно и пойдет в таблицу
        if link_of_video=="":
            sheet[f'E{i}']='-'
        else:
            sheet[f'E{i}']=link_of_video
        sheet[f'F{i}']=list_of_videos_names[j]
        sheet[f'G{i}'] = list_of_links[j]
        sheet[f'H{i}'] = list_of_channels[j]
        # увеличиваем счетчик на 1, чтобы перейти к следующим элементам списков имен, ссылок и каналов
        j+=1
        position+=1
    #     сохраняемся
    table_of_videos.save('table_of_videos.xlsx')
    return

# функция, которая анализирует данные бокового блока при посещении определенного канала
def get_channel(id_of_channel,quantity,driver):
    global link_of_video
    # # заходим на страницу канала и получаем куки
    driver.get(base_url+id_of_channel)
    # with open('cookies.txt', 'rb') as cookiesfile:
    #     cookies = pickle.load(cookiesfile)
    # for cookie in cookies:
    #     driver.add_cookie(cookie)
    # driver.refresh()
    # кликаем первое видео
    WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="video-title"]'))).click()
    # сохраняем в переменную ссылку на видео - она нам нужна и в первой функции
    link_of_video=driver.current_url
    # видео начинает проигрываться автоматически, ждем
    time.sleep(40)
    # получаем данные страницы
    html = driver.page_source
    # извлекаем данные о всех названиях, ссылках на видео и имен каналов из кода страницы
    soup = BS(html, "html.parser")
    list_of_videos_names = soup.find_all("span", {"id":"video-title","class":["style-scope ytd-compact-radio-renderer","style-scope ytd-compact-video-renderer"]})
    list_of_video_links=soup.find_all("a", {"class": ["yt-simple-endpoint style-scope ytd-compact-video-renderer","yt-simple-endpoint style-scope ytd-compact-radio-renderer"]})
    # list_of_video_links = soup.select(a[class='yt-simple-endpoint style-scope ytd-compact-video-renderer',a[class='yt-simple-endpoint style-scope ytd-compact-radio-renderer']
    html = driver.find_element_by_tag_name('html')
    html.send_keys(Keys.END)
    time.sleep(3)
    list_of_channels_all = driver.find_elements_by_xpath("// div[ @ id ='text-container'][@class ='style-scope ytd-channel-name']/yt-formatted-string[@ class ='style-scope ytd-channel-name'][@ id='text'][@ ellipsis-truncate=''][@ title='']")
    # list_of_channels_all = WebDriverWait(driver, 10).until(EC.presence_of_all_elements_located((By.XPATH,"// div[ @ id ='text-container'][@class ='style-scope ytd-channel-name']/yt-formatted-string[@ class ='style-scope ytd-channel-name'][@ id='text'][@ ellipsis-truncate=''][@ title='']")))
    # list_of_channels_all = soup.find_all("yt-formatted-string", {"id":"text","title":"","class":"style-scope ytd-channel-name","ellipsis-truncate":""})
    names=[]
    links=[]
    channels=[]
    # перебираем значения полученного массива видео, достаем оттуда имена и ссылки, помещаем их в новые списки
    for video in list_of_videos_names[0:20]:
        names.append(video.get('title'))
    for video in list_of_video_links[0:20]:
        links.append(base_url+video.get('href'))
    for video in list_of_channels_all:
        if (video.text)!='':
            channels.append(video.text)
    channels=channels[0:20]

    #     записываем в excel
    table_of_videos=load_workbook('table_of_videos.xlsx')
    # работаем на первом листе
    sheet = table_of_videos.active
    # ищем последнюю заполненную строку и переходим на следующую
    index=sheet.max_row+1
    #     заполняем 20 строк таблицы данными, которые ранее получили в ходе парсинга
    # заведем счетчик, чтобы идти по списку полученных данных
    j=0
    position=1
    for i in range(index,index+20):
        sheet[f'A{i}'] = 'random'
        sheet[f'B{i}'] = quantity
        sheet[f'C{i}'] = 'боковой блок'
        sheet[f'D{i}']= position
        sheet[f'E{i}'] = link_of_video   # ссылка на последнее просмотренное видео
        sheet[f'F{i}'] = names[j]
        sheet[f'G{i}'] = links[j]
        sheet[f'H{i}'] = channels[j]
        j+=1
        position+=1
    #     сохраняемся
    table_of_videos.save('table_of_videos.xlsx')

log_pass={}
# открываем файл
with open("auth.csv", encoding='utf-8') as r_file:
    reader_object = csv.reader(r_file, delimiter = ":")
    for row in reader_object:
        login,password=row[0],row[1]
        log_pass[login]=password

for key,value in log_pass.items():
    authorisation(driver,key,value)
    get_info_in_main_page(0,driver)
    for channel,count in channels.items():
        get_channel(channel,count,driver)
























