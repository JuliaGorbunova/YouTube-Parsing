from selenium import webdriver
import requests
from bs4 import BeautifulSoup as BS
import time
import xlwt
import openpyxl
from urllib.request import *
driver = webdriver.Chrome()
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from urllib.parse import urlparse
base_url = "https://www.youtube.com/"
# driver = webdriver.Chrome(r'C:\Users\Юлия\PycharmProjects\YouTube\chromedriver.exe')
# driver.implicitly_wait(5)
def get_info_in_main_page():
    driver.get(base_url)
    time.sleep(10)
    # логинимся
    # WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.XPATH,'//*[@id="buttons"]/ytd-button-renderer/a'))).click()
    # WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="identifierId"]'))).send_keys('marcovna-zin@yandex.ru')
    # WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="identifierNext"]/div/button/div[3]'))).click()
    # WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="password"]/div[1]/div/div[1]/input'))).send_keys('Qwerty12345@Qwerty')
    # WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="passwordNext"]/div/button/span'))).click()
    # получаем текст исходного кода страницы
    html = driver.page_source
    # извлекаем данные о всех названиях видео из кода страницы
    soup = BS(html, "html.parser")
    list_of_videos=soup.find_all("a",{"class":"yt-simple-endpoint focus-on-expand style-scope ytd-rich-grid-media"})
    list_of_channels_all=soup.find_all("a",{"class":"yt-simple-endpoint style-scope yt-formatted-string"})
    # создаем пустые списки, куда в дальнейшем поместим имена видео, ссылки на них, а также названия каналов
    list_of_videos_names=[]
    list_of_links = []
    list_of_channels=[]
    # перебираем значения полученного массива видео, достаем оттуда имена и ссылки, помещаем их в новые списки
    for video in list_of_videos[0:19]:
        list_of_links.append(base_url+video.get('href'))
        list_of_videos_names.append(video.get('title'))
    #  перебираем значения полученного массива каналов, достаем оттуда ссылки на каналы
    for item in list_of_channels_all[0:19]:
        list_of_channels.append(base_url+item.get('href'))
    # дальше работа с excel. Создаем таблицу (рабочую книгу)
    table_of_videos=openpyxl.Workbook()
    # делаем заглавие для первого листа
    table_of_videos.create_sheet(title='Статистика',index=0)
    # далее работаем на первом листе
    table_of_videos.save('table_of_videos.xlsx')
    sheet=table_of_videos['Статистика']
    # забиваем заголовки столбцов
    sheet['A1']='Аккаунт (рандомный айди)'
    sheet['B1']='Сколько видео из списка А1 посмотрел'
    sheet['C1']='Рекламный блок'
    sheet['D1']='Позиция'
    sheet['E1']='Ссылка на последнее просмотренное видео'
    sheet['F1']='Название'
    sheet['G1']='URL видео'
    sheet['H1']='URL канала'
    # забиваем значения первых двадцати строк
    for i in range(2,22):
        sheet[f'A{i}'] = 'random'
        sheet[f'B{i}']='0'
        sheet[f'C{i}']='главная страница'
        sheet[f'D{i}']=i-1
        sheet[f'E{i}']='-'
        sheet[f'F{i}']=list_of_videos_names[i-1]
        sheet[f'G{i}'] = list_of_links[i-1]
        sheet[f'H{i}'] = list_of_channels[i-1]
    #     сохраняемся
    table_of_videos.save('table_of_videos.xlsx')

get_info_in_main_page()

def get_channel(id_of_channel):
    index=None
    driver.get(base_url+id_of_channel)
    link=WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="video-title"]')))
    link.click()
    time.sleep(120)
    html = driver.page_source
    # извлекаем данные о всех названиях видео из кода страницы
    soup = BS(html, "html.parser")
    list_of_videos_names = soup.find_all("span", {"class": "style-scope ytd-compact-video-renderer"})
    list_of_video_links=soup.find_all("a", {"class": "yt-simple-endpoint style-scope ytd-compact-video-renderer"})
    # list_of_channels_all = soup.find_all("yt-formatted-string", {"id": "text"})
    names=[]
    links=[]
    channels=[]
    # перебираем значения полученного массива видео, достаем оттуда имена и ссылки, помещаем их в новые списки
    for video in list_of_videos_names[0:19]:
        names.append(video.get('title'))
    for video in list_of_video_links[0:19]:
        links.append(base_url+video.get('href'))
    # for video in list_of_channels[0:19]:
    #     channels.append(base_url + video.get('href'))
#     записываем в excel
    table_of_videos=openpyxl.load_workbook('table_of_videos.xlsx')
    # работаем на первом листе
    sheet=table_of_videos['Статистика']
#     ищем первую пустую строку, чтобы продолжить заполнять таблицу
    sheet=table_of_videos.active
    # для этого берем строки таблицы со 2 по 221 (т.к. в конце должно быть заполнено 220 строк)
    # и ищем первую пустую строку, строка перед которой НЕ пуста
    for i in range(1,221):
        if sheet[f'A{i}'].value=='' and sheet[f'A{i-1}']!='':
            index=i
    #     заполняем 20 строк таблицы данными, которые ранее получили в ходе парсинга
    for i in range(index,index+20):
        sheet[f'A{i}'] = 'random'
        sheet[f'B{i}'] = '1'
        sheet[f'C{i}'] = 'боковой блок'
        # sheet[f'D{i}'] = i - 1
        sheet[f'E{i}'] = '-'   # ссылка на последнее просмотренное видео
        # sheet[f'F{i}'] = list_of_videos_names[i - 1]
        # sheet[f'G{i}'] = list_of_links[i - 1]
        # sheet[f'H{i}'] = list_of_channels[i - 1]

    get_info_in_main_page()

get_channel('channel/UCMCgOm8GZkHp8zJ6l7_hIuA')





























